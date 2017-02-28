import time
import openpyxl
import requests
from bs4 import BeautifulSoup

save_num = []

workbook = openpyxl.load_workbook('test3.xlsx')
sheet_name = workbook.get_sheet_names()[0]
sheet = workbook.get_sheet_by_name(sheet_name)
num_rows = sheet.max_row
num_cols = sheet.max_column

init = []
for row in range(num_rows):
  r = {}
  for col in range(num_cols):
    if col == 0:
      r['author'] = sheet.cell(row=row + 1, column=col + 1).value
    elif col == 1:
      r['book'] = sheet.cell(row=row + 1, column=col + 1).value
  init.append(r)

print(init)
cc = 1

for i in init:
  i['author'] = i['author'].replace(" ", "")
  author = i['author']
  book = i['book']

  data1 = {'action': 'result', 'autor': author}
  result = requests.post("http://www.unforbi.com.ar/cutteren/index.php",
                         data=data1)
  soup1 = BeautifulSoup(result.content, 'html.parser')
  num = str(soup1.find('strong')).replace("<strong>", "").replace("</strong>","")
  time.sleep(2)

  firstbookname = book[0]
  number = str(num) + str(firstbookname)
  save_num.append(number)

  cnt = save_num.count(number)

  if cnt > 1:
    number = number + ' c' + str(cnt)

  sheet.cell(row=cc, column=3, value=number)
  cc += 1

workbook.save("test3.xlsx")


