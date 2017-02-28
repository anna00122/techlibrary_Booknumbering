from hangul_utils import split_syllables, join_jamos
import openpyxl


def all_num2(ip):
  if ip == 'ㅏ':
    ip_value = '2'
  elif (ip == 'ㅐ') or (ip == 'ㅑ') or (ip == 'ㅒ'):
    ip_value = '3'
  elif (ip == 'ㅓ') or (ip == 'ㅔ') or (ip == 'ㅕ') or (ip == 'ㅖ'):
    ip_value = '4'
  elif (ip == 'ㅗ') or (ip == 'ㅘ') or (ip == 'ㅙ') or (ip == 'ㅚ') or (ip == 'ㅛ'):
    ip_value = '5'
  elif (ip == 'ㅜ') or (ip == 'ㅝ') or (ip == 'ㅞ') or (ip == 'ㅟ') or (ip == 'ㅠ'):
    ip_value = '6'
  elif (ip == 'ㅡ') or (ip == 'ㅢ'):
    ip_value = '7'
  elif ip == 'ㅣ':
    ip_value = '8'
  return ip_value


def ch_num2(ip):
  if (ip == 'ㅏ') or (ip == 'ㅐ') or (ip == 'ㅑ') or (ip == 'ㅒ'):
    ip_value = '2'
  elif (ip == 'ㅓ') or (ip == 'ㅔ') or (ip == 'ㅕ') or (ip == 'ㅖ'):
    ip_value = '3'
  elif (ip == 'ㅗ') or (ip == 'ㅘ') or (ip == 'ㅙ') or (ip == 'ㅚ') or (ip == 'ㅛ'):
    ip_value = '4'
  elif (ip == 'ㅜ') or (ip == 'ㅝ') or (ip == 'ㅞ') or (ip == 'ㅟ') or (
    ip == 'ㅠ') or (ip == 'ㅡ') or (ip == 'ㅢ'):
    ip_value = '5'
  elif ip == 'ㅣ':
    ip_value = '6'
  return ip_value



save_num = []  ##청구기호저장


workbook = openpyxl.load_workbook('list.xlsx')
sheet_name = workbook.get_sheet_names()[0]
sheet = workbook.get_sheet_by_name(sheet_name)
num_rows = sheet.max_row
num_cols = sheet.max_column


init = []
for row in range(num_rows):
  r = {}
  for col in range(num_cols):
    if col == 4:
      r['author'] = sheet.cell(row=row+1, column=col+1).value
    elif col == 2:
      r['book'] = sheet.cell(row=row+1, column=col+1).value
  init.append(r)



print(init)



cc = 1

for i in init:
  i['author'] = i['author'].replace(" ", "")
  author = i['author']
  book = i['book']

  jamo = split_syllables(author[1])
  jamo_list = list(jamo)

  if jamo_list[0] == 'ㅊ':
    num1 = '8'
    num2 = ch_num2(jamo_list[1])

  elif (jamo_list[0] == 'ㄱ') or (jamo_list[0] == 'ㄲ'):
    num1 = '1'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㄴ':
    num1 = '19'
    num2 = all_num2(jamo_list[1])

  elif (jamo_list[0] == 'ㄷ') or (jamo_list[0] == 'ㄸ'):
    num1 = '2'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㄹ':
    num1 = '29'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㅁ':
    num1 = '3'
    num2 = all_num2(jamo_list[1])

  elif (jamo_list[0] == 'ㅂ') or (jamo_list[0] == 'ㅃ'):
    num1 = '4'
    num2 = all_num2(jamo_list[1])

  elif (jamo_list[0] == 'ㅅ') or (jamo_list[0] == 'ㅆ'):
    num1 = '5'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㅇ':
    num1 = '6'
    num2 = all_num2(jamo_list[1])

  elif (jamo_list[0] == 'ㅈ') or (jamo_list[0] == 'ㅉ'):
    num1 = '7'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㅋ':
    num1 = '87'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㅌ':
    num1 = '88'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㅍ':
    num1 = '89'
    num2 = all_num2(jamo_list[1])

  elif jamo_list[0] == 'ㅎ':
    num1 = '9'
    num2 = all_num2(jamo_list[1])

  first = author[0]
  second = num1 + num2

  bookname_jamo = split_syllables(book)
  bookname_jamo_list = list(bookname_jamo)
  bookname_first = bookname_jamo_list[0]

  num = first + second + bookname_first
  save_num.append(num)
  cnt = save_num.count(num)

  if cnt > 1:
    num = num + 'c' + str(cnt)

  sheet.cell(row=cc, column=7, value=num)

  cc += 1

workbook.save("list.xlsx")

